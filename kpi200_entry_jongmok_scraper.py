import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

KPI200_ENTRY_URL = "https://finance.naver.com/sise/entryJongmok.naver?type=KPI200"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
}


def build_page_url(page: int) -> str:
    return f"{KPI200_ENTRY_URL}&page={page}"


def parse_table_rows(table, headers: list[str]) -> list[dict[str, str]]:
    rows = []
    for tr in table.find_all("tr"):
        cells = tr.find_all("td")
        if not cells or len(cells) != len(headers):
            continue

        values = [cell.get_text(strip=True) for cell in cells]
        if not any(values):
            continue

        row = {headers[i]: values[i] for i in range(len(headers))}
        rows.append(row)

    return rows


def fetch_kpi200_entry_rows(page_count: int = 20) -> list[dict[str, str]]:
    """Fetch KPI200 편입종목상위 data from Naver Finance across multiple pages."""
    all_rows: list[dict[str, str]] = []
    headers: list[str] | None = None

    for page in range(1, page_count + 1):
        response = requests.get(build_page_url(page), headers=HEADERS)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, "html.parser")
        table = soup.find("table", class_="type_1")
        if table is None:
            raise RuntimeError(f"Cannot find KPI200 entry table on page {page}.")

        if headers is None:
            header_cells = table.find_all("th")
            headers = [cell.get_text(strip=True) for cell in header_cells]
            if not headers:
                raise RuntimeError("Cannot parse table headers from KPI200 entry page.")

        page_rows = parse_table_rows(table, headers)
        all_rows.extend(page_rows)
        print(f"Page {page}: fetched {len(page_rows)} rows")

    return all_rows


def save_rows_to_excel(rows: list[dict[str, str]], filename: str = "kospi200.xlsx") -> str:
    if not rows:
        raise ValueError("No rows to save.")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "KOSPI200"

    headers = list(rows[0].keys())
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])

    for col_index in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_index)
        worksheet.column_dimensions[column_letter].width = 18

    workbook.save(filename)
    return filename


if __name__ == "__main__":
    try:
        entry_rows = fetch_kpi200_entry_rows(page_count=20)
        if not entry_rows:
            print("No entry rows found.")
        else:
            output_file = save_rows_to_excel(entry_rows, "kospi200.xlsx")
            print(f"Fetched {len(entry_rows)} KPI200 편입종목 rows across 20 pages.")
            print(f"Saved crawled data to {output_file}")
    except Exception as exc:
        print("Error fetching KPI200 entry data:", exc)
